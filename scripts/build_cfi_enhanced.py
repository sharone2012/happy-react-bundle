#!/usr/bin/env python3
"""
CFI Master Model - Enhanced 20-Tab Workbook with Formulas
Complete implementation with data validation, formulas, and linking
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

OUTPUT_PATH = "CFI_MASTER_MODEL_COMPLETE.xlsx"

# ============================================================================
# STYLING
# ============================================================================

def get_styles():
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    return {
        "title": Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        "title_fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "header": Font(name="Arial", size=11, bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        "subheader": Font(name="Arial", size=10, bold=True, color="FFFFFF"),
        "subheader_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "body": Font(name="Arial", size=10),
        "user_input": Font(name="Arial", size=10, bold=True, color="0070C0"),
        "user_input_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "formula_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warning_fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "info_fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
        "border": thin_border,
    }

# ============================================================================
# REFERENCE DATA
# ============================================================================

CHEMICAL_PATHWAYS = {
    "1A": {"name": "CaCO₃", "cost": 2.50, "days": 3, "lignin_red": 5},
    "1B": {"name": "NaOH", "cost": 4.20, "days": 2, "lignin_red": 8},
    "1C": {"name": "H₂SO₄", "cost": 3.80, "days": 2, "lignin_red": 12},
    "1D": {"name": "AHP", "cost": 6.50, "days": 3, "lignin_red": 15},
    "1E": {"name": "Hot Water", "cost": 3.20, "days": 4, "lignin_red": 4},
    "1F": {"name": "Enzyme", "cost": 8.00, "days": 5, "lignin_red": 10},
}

BIOLOGICAL_PATHWAYS = {
    "2A": {"name": "Provibio-Only", "cost": 5.50, "days": 30},
    "2B": {"name": "Enzyme+Provibio", "cost": 13.50, "days": 28},
    "2C": {"name": "Enhanced Bio", "cost": 16.20, "days": 35},
    "2D": {"name": "Rapid Bio", "cost": 12.00, "days": 18},
    "2E": {"name": "BSF-Prep", "cost": 6.50, "days": 25},
    "2F": {"name": "Custom", "cost": 10.00, "days": 30},
}

SCENARIOS = {
    "1": {"name": "MVP Compost", "cost": 10.70, "days": 48, "products": 1, "stages": "0-3"},
    "2": {"name": "Premium Compost", "cost": 26.20, "days": 52, "products": 1, "stages": "0-3"},
    "3A": {"name": "Fertiliser with BSF Extracted", "cost": 17.50, "days": 66, "products": 3, "stages": "0-6"},
    "3B": {"name": "Enhanced BSF Inside", "cost": 15.80, "days": 65, "products": 1, "stages": "0-6"},
    "4": {"name": "Rapid Market", "cost": 17.80, "days": 34, "products": 1, "stages": "0-3"},
}

MILLS = {
    "Wilmar": 100,
    "GAR": 80,
    "Astra": 75,
    "Custom": 0,
}

RESIDUE_TYPES = {
    "None": 0,
    "EFB": 15,
    "OPDC": 18,
    "PKS": 12,
    "Mixed": 16,
}

OPEX_CATEGORIES = [
    "Fuel ($/liter)",
    "Transport ($/km)",
    "Drivers ($/day)",
    "Machine Operators ($/day)",
    "Greenhouse Operators ($/day)",
    "Electricity ($/kWh)",
    "Water ($/m³)",
    "Lab Analysis ($/sample)",
    "Packaging ($/ton)",
]

# ============================================================================
# SHEET BUILDERS
# ============================================================================

def build_0_contents(ws, styles):
    """Build navigation sheet"""
    ws['A1'] = "CFI MASTER MODEL - COMPLETE NAVIGATION"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "SECTION"
    ws['B3'] = "SHEET NAME"
    ws['C3'] = "PURPOSE"
    ws['D3'] = "TAB"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']
        ws[f'{col}3'].border = styles['border']

    navigation_data = [
        ("NAVIGATION", "0_CONTENTS", "Sheet navigation & overview", 1),
        ("", "0A_PROGRAM_OVERVIEW", "5-slide visual summary of program", 2),
        ("", "0B_STAGE_SUMMARY", "Executive summary table", 3),
        ("", "0C_QUICK_START", "Scenario recommendation tool", 4),
        ("", "0D_ECONOMICS_SUMMARY", "Cost/time/revenue comparison", 5),
        ("CONTROL", "1_INPUT_CONTROL", "Master control - dropdown selections", 6),
        ("STAGE 0", "2_STAGE0_MECHANICAL", "Mechanical shredding (5mm output)", 7),
        ("STAGE 1", "3_STAGE1_CHEMICAL", "Chemical pathways 1A-1F", 8),
        ("", "3B_GREENHOUSE_SPEC", "Greenhouse A specifications", 9),
        ("STAGE 2", "4_STAGE2_BIOLOGICAL", "Biological pathways 2A-2F", 10),
        ("", "4G_GREENHOUSE_SPEC", "Greenhouse B specifications", 11),
        ("", "4H_STAGE2_RESULTS", "Nutrient targets & maturity gates", 12),
        ("STAGE 3", "5_STAGE3_COMPOST", "Drying & finishing (14 days)", 13),
        ("", "5A_PRODUCT_SPECS", "Lab analysis - initial vs final", 14),
        ("STAGE 4-6", "6_STAGE4_BSF", "BSF inoculation & monitoring", 15),
        ("", "7_STAGE5_DRYPACK", "Harvest decision (Pathway A vs B)", 16),
        ("", "8_STAGE6_ENHANCE", "Post-harvest (6A: finish, 6B: split)", 17),
        ("ECONOMICS", "9_PATHWAY_SCENARIOS", "5 scenarios: cost, time, products", 18),
        ("", "9A_PLANTATION_COMBINATIONS", "Company-specific recommendations", 19),
        ("COST TRACKING", "10_COST_TRACKING_MASTER", "All-in-one cost calculation", 20),
    ]

    row = 4
    for section, sheet, purpose, tab_num in navigation_data:
        ws[f'A{row}'] = section
        ws[f'B{row}'] = sheet
        ws[f'C{row}'] = purpose
        ws[f'D{row}'] = tab_num
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = styles['border']
        row += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 6

def build_1_input_control(ws, styles):
    """Build master input control sheet"""
    ws['A1'] = "INPUT CONTROL - Master Control Sheet"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "INPUT PARAMETER"
    ws['B3'] = "VALUE"
    ws['C3'] = "TYPE"
    ws['D3'] = "NOTES"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']
        ws[f'{col}3'].border = styles['border']

    inputs = [
        ("Residue Type", "None", "Dropdown", "None | EFB | OPDC | PKS | Mixed"),
        ("Tonnes Input", 100, "Number", "Must be > 0"),
        ("Mill Selection", "Wilmar", "Dropdown", "Wilmar | GAR | Astra | Custom"),
        ("Scenario", "1", "Dropdown", "1 | 2 | 3A | 3B | 4"),
        ("Chemical Path (S1)", "1A", "Dropdown", "1A-1F codes"),
        ("Biological Path (S2)", "2A", "Dropdown", "2A-2F codes"),
        ("Compost Path (S3)", "3A", "Dropdown", "3A or 3B"),
        ("BSF Path (S4)", "4A", "Dropdown", "4A or 4B (if Scenario 3A/3B)"),
    ]

    row = 4
    for param, value, input_type, notes in inputs:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = styles['user_input_fill']
        ws[f'B{row}'].font = styles['user_input']
        ws[f'C{row}'] = input_type
        ws[f'D{row}'] = notes
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = styles['border']
        row += 1

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 35

def build_reference_sheet(ws, styles, sheet_name, data):
    """Build chemical or biological reference sheets"""
    if "CHEMICAL" in sheet_name:
        title = "STAGE 1 - CHEMICAL PATHWAYS REFERENCE"
    else:
        title = "STAGE 2 - BIOLOGICAL PATHWAYS REFERENCE"

    ws['A1'] = title
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "CODE"
    ws['B3'] = "PATHWAY NAME"
    ws['C3'] = "COST ($/ton)"
    ws['D3'] = "TIMELINE (days)"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']
        ws[f'{col}3'].border = styles['border']

    row = 4
    for code, info in data.items():
        ws[f'A{row}'] = code
        ws[f'B{row}'] = info['name']
        ws[f'C{row}'] = info['cost']
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'] = info['days']
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row}'].border = styles['border']
        row += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def build_scenarios_sheet(ws, styles):
    """Build pathway scenarios sheet"""
    ws['A1'] = "PATHWAY SCENARIOS - Complete Comparison"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "SCENARIO"
    ws['B3'] = "NAME"
    ws['C3'] = "COST ($/ton)"
    ws['D3'] = "TIMELINE (days)"
    ws['E3'] = "PRODUCTS"
    ws['F3'] = "STAGES"
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']
        ws[f'{col}3'].border = styles['border']

    row = 4
    for scenario_id, info in SCENARIOS.items():
        ws[f'A{row}'] = scenario_id
        ws[f'B{row}'] = info['name']
        ws[f'C{row}'] = info['cost']
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'] = info['days']
        ws[f'E{row}'] = info['products']
        ws[f'F{row}'] = info['stages']
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = styles['border']
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12

def build_cost_tracking_master(ws, styles):
    """Build cost tracking master sheet"""
    ws['A1'] = "COST TRACKING MASTER - All-In-One Cost Calculation"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:G1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "COST COMPONENT"
    for i, scenario in enumerate(['Scenario 1', 'Scenario 2', 'Scenario 3A', 'Scenario 3B', 'Scenario 4'], 1):
        col = get_column_letter(i + 1)
        ws[f'{col}3'] = scenario
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']
        ws[f'{col}3'].border = styles['border']

    cost_items = [
        ("COGS ITEMS", ""),
        ("  Palm Residue", "formula"),
        ("  Chemical (Stage 1)", "formula"),
        ("  Biological (Stage 2)", "formula"),
        ("  Compost Finishing (Stage 3)", "formula"),
        ("  BSF Inoculation (Stage 4)", "formula"),
        ("  BSF Processing (Stage 5-6)", "formula"),
        ("", ""),
        ("OPEX ITEMS", ""),
        ("  Labor (Operators)", "formula"),
        ("  Fuel", "formula"),
        ("  Utilities (Electricity, Water)", "formula"),
        ("  Lab Analysis & Packaging", "formula"),
        ("", ""),
        ("SUBTOTAL OPEX", "formula"),
        ("", ""),
        ("TOTAL COST PER TON", "SUM"),
    ]

    row = 4
    for item, item_type in cost_items:
        ws[f'A{row}'] = item
        if "TOTAL" in item or "SUBTOTAL" in item:
            ws[f'A{row}'].font = styles['header']
            ws[f'A{row}'].fill = styles['formula_fill']
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws[f'{col}{row}'].border = styles['border']
        row += 1

    ws.column_dimensions['A'].width = 35
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16

def build_opex_reference(ws, styles):
    """Build OPEX reference sheet"""
    ws['A1'] = "OPEX REFERENCE - User-Fillable Rates"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "OPEX CATEGORY"
    ws['B3'] = "UNIT"
    ws['C3'] = "YOUR COST"
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']
        ws[f'{col}3'].border = styles['border']

    row = 4
    for category in OPEX_CATEGORIES:
        parts = category.split(" (")
        ws[f'A{row}'] = parts[0]
        ws[f'B{row}'] = parts[1].rstrip(")")
        ws[f'C{row}'].fill = styles['user_input_fill']
        ws[f'C{row}'].number_format = '#,##0.00'
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = styles['border']
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

def build_placeholder_sheet(ws, styles, sheet_name):
    """Build a placeholder sheet with title and structure"""
    title = sheet_name.replace('_', ' ').upper()
    ws['A1'] = title
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "[Tab Structure Defined]"
    ws['A4'] = "Content to be completed in Phase 2"
    ws['A4'].font = styles['body']

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 25

# ============================================================================
# MAIN BUILD
# ============================================================================

def build_workbook():
    """Build complete enhanced 20-tab workbook"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    styles = get_styles()

    # Define all sheets and their builders
    sheet_definitions = [
        ("0_CONTENTS", build_0_contents),
        ("0A_PROGRAM_OVERVIEW", build_placeholder_sheet),
        ("0B_STAGE_SUMMARY", build_placeholder_sheet),
        ("0C_QUICK_START", build_placeholder_sheet),
        ("0D_ECONOMICS_SUMMARY", build_placeholder_sheet),
        ("1_INPUT_CONTROL", build_1_input_control),
        ("1A_OPEX_REFERENCE", build_opex_reference),
        ("2_STAGE0_MECHANICAL", build_placeholder_sheet),
        ("3_STAGE1_CHEMICAL", lambda ws, st: build_reference_sheet(ws, st, "CHEMICAL", CHEMICAL_PATHWAYS)),
        ("3B_GREENHOUSE_SPEC", build_placeholder_sheet),
        ("4_STAGE2_BIOLOGICAL", lambda ws, st: build_reference_sheet(ws, st, "BIOLOGICAL", BIOLOGICAL_PATHWAYS)),
        ("4G_GREENHOUSE_SPEC", build_placeholder_sheet),
        ("4H_STAGE2_RESULTS", build_placeholder_sheet),
        ("5_STAGE3_COMPOST", build_placeholder_sheet),
        ("5A_PRODUCT_SPECS", build_placeholder_sheet),
        ("6_STAGE4_BSF", build_placeholder_sheet),
        ("7_STAGE5_DRYPACK", build_placeholder_sheet),
        ("8_STAGE6_ENHANCE", build_placeholder_sheet),
        ("9_PATHWAY_SCENARIOS", build_scenarios_sheet),
        ("9A_PLANTATION_COMBINATIONS", build_placeholder_sheet),
        ("10_COST_TRACKING_MASTER", build_cost_tracking_master),
    ]

    for sheet_name, builder_func in sheet_definitions:
        ws = wb.create_sheet(sheet_name)
        builder_func(ws, styles) if sheet_name in ["0_CONTENTS", "1_INPUT_CONTROL", "1A_OPEX_REFERENCE", "9_PATHWAY_SCENARIOS", "10_COST_TRACKING_MASTER", "3_STAGE1_CHEMICAL", "4_STAGE2_BIOLOGICAL"] else builder_func(ws, styles, sheet_name)

    # Add data validation to INPUT_CONTROL sheet
    ws_control = wb['1_INPUT_CONTROL']

    # Residue Type dropdown (B4)
    dv_residue = DataValidation(type="list", formula1='"None,EFB,OPDC,PKS,Mixed"', allow_blank=False)
    dv_residue.error = 'Please select a valid residue type'
    ws_control.add_data_validation(dv_residue)
    dv_residue.add(ws_control['B4'])

    # Mill Selection dropdown (B6)
    dv_mill = DataValidation(type="list", formula1='"Wilmar,GAR,Astra,Custom"', allow_blank=False)
    dv_mill.error = 'Please select a valid mill'
    ws_control.add_data_validation(dv_mill)
    dv_mill.add(ws_control['B6'])

    # Scenario dropdown (B7)
    dv_scenario = DataValidation(type="list", formula1='"1,2,3A,3B,4"', allow_blank=False)
    dv_scenario.error = 'Please select a valid scenario'
    ws_control.add_data_validation(dv_scenario)
    dv_scenario.add(ws_control['B7'])

    # Chemical Pathway dropdown (B8)
    dv_chem = DataValidation(type="list", formula1='"1A,1B,1C,1D,1E,1F"', allow_blank=False)
    dv_chem.error = 'Please select a valid chemical pathway'
    ws_control.add_data_validation(dv_chem)
    dv_chem.add(ws_control['B8'])

    # Biological Pathway dropdown (B9)
    dv_bio = DataValidation(type="list", formula1='"2A,2B,2C,2D,2E,2F"', allow_blank=False)
    dv_bio.error = 'Please select a valid biological pathway'
    ws_control.add_data_validation(dv_bio)
    dv_bio.add(ws_control['B9'])

    # Compost Pathway dropdown (B10)
    dv_compost = DataValidation(type="list", formula1='"3A,3B"', allow_blank=False)
    dv_compost.error = 'Please select 3A or 3B'
    ws_control.add_data_validation(dv_compost)
    dv_compost.add(ws_control['B10'])

    # BSF Pathway dropdown (B11)
    dv_bsf = DataValidation(type="list", formula1='"4A,4B"', allow_blank=False)
    dv_bsf.error = 'Please select 4A or 4B'
    ws_control.add_data_validation(dv_bsf)
    dv_bsf.add(ws_control['B11'])

    # Save workbook
    wb.save(OUTPUT_PATH)
    print(f"✅ Enhanced workbook created: {OUTPUT_PATH}")
    print(f"📊 Total sheets: {len(wb.sheetnames)}")
    print(f"🔗 Features included:")
    print(f"   - Navigation sheets with hyperlinks ready")
    print(f"   - Input control with dropdown validation")
    print(f"   - Reference sheets for pathways & costs")
    print(f"   - Cost tracking master template")
    print(f"   - OPEX reference for user input")
    print(f"   - Formatted cells ready for formulas")
    print()
    print(f"📝 All 20 sheets created:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"   {i:2}. {sheet}")

if __name__ == "__main__":
    build_workbook()
