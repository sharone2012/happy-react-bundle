#!/usr/bin/env python3
"""
CFI Master Model - Optimized 20-Tab Workbook Builder
Circular Farming Indonesia - Complete Build from Specification
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

OUTPUT_PATH = "CFI_MASTER_MODEL_20TABS.xlsx"

# ============================================================================
# STYLING DEFINITIONS
# ============================================================================

def get_styles():
    return {
        "title": Font(name="Arial", size=14, bold=True, color="FFFFFF"),
        "title_fill": PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid"),
        "header": Font(name="Arial", size=11, bold=True, color="FFFFFF"),
        "header_fill": PatternFill(start_color="366092", end_color="366092", fill_type="solid"),
        "subheader": Font(name="Arial", size=10, bold=True, color="FFFFFF"),
        "subheader_fill": PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid"),
        "body": Font(name="Arial", size=10),
        "user_input": Font(name="Arial", size=10, bold=True, color="0070C0"),
        "user_input_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "formula_fill": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        "warning_fill": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
        "info_fill": PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid"),
        "center": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "left": Alignment(horizontal="left", vertical="center", wrap_text=True),
        "right": Alignment(horizontal="right", vertical="center"),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
    }

# ============================================================================
# REFERENCE DATA
# ============================================================================

CHEMICAL_PATHWAYS = {
    "1A": {"name": "CaCO₃", "cost": 2.50, "days": 3, "lignin_reduction": 5},
    "1B": {"name": "NaOH", "cost": 4.20, "days": 2, "lignin_reduction": 8},
    "1C": {"name": "H₂SO₄", "cost": 3.80, "days": 2, "lignin_reduction": 12},
    "1D": {"name": "AHP", "cost": 6.50, "days": 3, "lignin_reduction": 15},
    "1E": {"name": "Hot Water", "cost": 3.20, "days": 4, "lignin_reduction": 4},
    "1F": {"name": "Enzyme", "cost": 8.00, "days": 5, "lignin_reduction": 10},
}

BIOLOGICAL_PATHWAYS = {
    "2A": {"name": "Provibio-Only", "cost": 5.50, "days": 30},
    "2B": {"name": "Enzyme+Provibio", "cost": 13.50, "days": 28},
    "2C": {"name": "Enhanced Bio", "cost": 16.20, "days": 35},
    "2D": {"name": "Rapid Bio", "cost": 12.00, "days": 18},
    "2E": {"name": "BSF-Prep", "cost": 6.50, "days": 25},
    "2F": {"name": "Custom", "cost": 10.00, "days": 30},
}

SCENARIOS = {
    "1": {"name": "MVP Compost", "cost": 10.70, "days": 48, "products": 1},
    "2": {"name": "Premium Compost", "cost": 26.20, "days": 52, "products": 1},
    "3A": {"name": "Fertiliser with BSF Extracted", "cost": 17.50, "days": 66, "products": 3},
    "3B": {"name": "Enhanced Fertiliser with BSF Inside", "cost": 15.80, "days": 65, "products": 1},
    "4": {"name": "Rapid Market", "cost": 17.80, "days": 34, "products": 1},
}

# ============================================================================
# SHEET BUILDERS
# ============================================================================

def build_0_contents(ws, styles):
    """Navigation sheet"""
    ws['A1'] = "CFI MASTER MODEL - NAVIGATION"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = styles['center']

    ws.row_dimensions[1].height = 25

    row = 3
    ws[f'A{row}'] = "SECTION"
    ws[f'B{row}'] = "SHEET NAME"
    ws[f'C{row}'] = "PURPOSE"
    ws[f'D{row}'] = "TAB #"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}{row}'].font = styles['header']
        ws[f'{col}{row}'].fill = styles['header_fill']

    row = 4
    sections = [
        ("NAVIGATION", "0_CONTENTS", "This sheet - quick navigation", 1),
        ("", "0A_PROGRAM_OVERVIEW", "5-slide visual summary", 2),
        ("", "0B_STAGE_SUMMARY", "Executive stage table", 3),
        ("", "0C_QUICK_START", "Decision tree for scenarios", 4),
        ("", "0D_ECONOMICS_SUMMARY", "Cost/time/profit comparison", 5),
        ("CONTROL", "1_INPUT_CONTROL", "Master control sheet with dropdowns", 6),
        ("STAGE 0", "2_STAGE0_MECHANICAL", "Shredding specs & output", 7),
        ("STAGE 1", "3_STAGE1_CHEMICAL", "Chemical pathways 1A-1F", 8),
        ("", "3B_GREENHOUSE_SPEC", "Greenhouse A specs", 9),
        ("STAGE 2", "4_STAGE2_BIOLOGICAL", "Biological pathways 2A-2F", 10),
        ("", "4G_GREENHOUSE_SPEC", "Greenhouse B specs", 11),
        ("", "4H_STAGE2_RESULTS", "Target nutrients & maturity", 12),
        ("STAGE 3", "5_STAGE3_COMPOST", "Drying & finishing", 13),
        ("", "5A_PRODUCT_SPECS", "Lab analysis & quality gates", 14),
        ("STAGE 4-6", "6_STAGE4_BSF", "BSF inoculation & monitoring", 15),
        ("", "7_STAGE5_DRYPACK", "Harvest & termination decision", 16),
        ("", "8_STAGE6_ENHANCE", "Post-harvest processing (A & B)", 17),
        ("ECONOMICS", "9_PATHWAY_SCENARIOS", "5 scenarios with full breakdown", 18),
        ("", "9A_PLANTATION_COMBINATIONS", "Company-specific combos", 19),
        ("COST TRACKING", "10_COST_TRACKING_MASTER", "All-in-one cost calculation", 20),
    ]

    for section, sheet_name, purpose, tab_num in sections:
        ws[f'A{row}'] = section
        ws[f'B{row}'] = sheet_name
        ws[f'C{row}'] = purpose
        ws[f'D{row}'] = tab_num
        row += 1

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 8

def build_1_input_control(ws, styles):
    """Master input control sheet"""
    ws['A1'] = "INPUT CONTROL - Master Control Sheet"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:C1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    # User inputs
    ws['A3'] = "PARAMETER"
    ws['B3'] = "VALUE"
    ws['C3'] = "NOTE"
    for col in ['A', 'B', 'C']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']

    inputs = [
        ("Residue Type", "EFB", "None, EFB, OPDC, PKS, Mixed"),
        ("Tonnes Input", 100, "Enter number > 0"),
        ("Mill Selection", "Wilmar", "Wilmar, GAR, Astra, Custom"),
        ("Scenario", 1, "1, 2, 3A, 3B, or 4"),
        ("Chemical Pathway (Stage 1)", "1A", "1A-1F"),
        ("Biological Pathway (Stage 2)", "2A", "2A-2F"),
        ("Compost Pathway (Stage 3)", "3A", "3A or 3B"),
        ("BSF Pathway (Stage 4)", "4A", "4A or 4B (if scenario = 3A/3B)"),
    ]

    row = 4
    for param, value, note in inputs:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = value
        ws[f'B{row}'].fill = styles['user_input_fill']
        ws[f'B{row}'].font = styles['user_input']
        ws[f'C{row}'] = note
        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 40

def build_reference_sheets(ws, styles, sheet_name):
    """Build chemical and biological reference sheets"""
    if "CHEMICAL" in sheet_name.upper():
        data = CHEMICAL_PATHWAYS
        ws['A1'] = "STAGE 1 - CHEMICAL PATHWAYS"
    else:
        data = BIOLOGICAL_PATHWAYS
        ws['A1'] = "STAGE 2 - BIOLOGICAL PATHWAYS"

    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "CODE"
    ws['B3'] = "NAME"
    ws['C3'] = "COST ($/ton)"
    ws['D3'] = "DAYS"
    for col in ['A', 'B', 'C', 'D']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']

    row = 4
    for code, info in data.items():
        ws[f'A{row}'] = code
        ws[f'B{row}'] = info['name']
        ws[f'C{row}'] = info['cost']
        ws[f'D{row}'] = info['days']
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12

def build_scenarios_sheet(ws, styles):
    """Scenario comparison sheet"""
    ws['A1'] = "PATHWAY SCENARIOS - Complete Comparison"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:E1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "SCENARIO"
    ws['B3'] = "NAME"
    ws['C3'] = "COST ($/ton)"
    ws['D3'] = "TIMELINE (days)"
    ws['E3'] = "FINAL PRODUCTS"
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']

    row = 4
    for scenario_id, info in SCENARIOS.items():
        ws[f'A{row}'] = scenario_id
        ws[f'B{row}'] = info['name']
        ws[f'C{row}'] = info['cost']
        ws[f'D{row}'] = info['days']
        ws[f'E{row}'] = info['products']
        row += 1

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

def build_cost_tracking_master(ws, styles):
    """Cost tracking master sheet"""
    ws['A1'] = "COST TRACKING MASTER - All-in-One Calculation"
    ws['A1'].font = styles['title']
    ws['A1'].fill = styles['title_fill']
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = styles['center']
    ws.row_dimensions[1].height = 25

    ws['A3'] = "COST ITEM"
    for i, scenario in enumerate(['1', '2', '3A', '3B', '4'], 1):
        col = get_column_letter(i + 1)
        ws[f'{col}3'] = f"Scenario {scenario}"
        ws[f'{col}3'].font = styles['header']
        ws[f'{col}3'].fill = styles['header_fill']

    cost_items = [
        "Palm Residue",
        "Chemical (Stage 1)",
        "Biological (Stage 2)",
        "Compost (Stage 3)",
        "BSF Inoculation (Stage 4)",
        "Harvest/Termination (Stage 5)",
        "Finishing (Stage 6)",
        "OPEX - Fuel",
        "OPEX - Labor",
        "OPEX - Utilities",
        "Total Cost/ton"
    ]

    row = 4
    for item in cost_items:
        ws[f'A{row}'] = item
        if row == 14:  # Total row
            ws[f'A{row}'].font = styles['header']
            ws[f'A{row}'].fill = styles['formula_fill']
        row += 1

    ws.column_dimensions['A'].width = 25
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 15

# ============================================================================
# MAIN BUILD FUNCTION
# ============================================================================

def build_workbook():
    """Build complete 20-tab workbook"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    styles = get_styles()

    # Define sheets
    sheets_to_create = [
        ("0_CONTENTS", "build_0_contents"),
        ("0A_PROGRAM_OVERVIEW", "basic"),
        ("0B_STAGE_SUMMARY", "basic"),
        ("0C_QUICK_START", "basic"),
        ("0D_ECONOMICS_SUMMARY", "basic"),
        ("1_INPUT_CONTROL", "build_1_input_control"),
        ("2_STAGE0_MECHANICAL", "basic"),
        ("3_STAGE1_CHEMICAL", "build_reference_sheets"),
        ("3B_GREENHOUSE_SPEC", "basic"),
        ("4_STAGE2_BIOLOGICAL", "build_reference_sheets"),
        ("4G_GREENHOUSE_SPEC", "basic"),
        ("4H_STAGE2_RESULTS", "basic"),
        ("5_STAGE3_COMPOST", "basic"),
        ("5A_PRODUCT_SPECS", "basic"),
        ("6_STAGE4_BSF", "basic"),
        ("7_STAGE5_DRYPACK", "basic"),
        ("8_STAGE6_ENHANCE", "basic"),
        ("9_PATHWAY_SCENARIOS", "build_scenarios_sheet"),
        ("9A_PLANTATION_COMBINATIONS", "basic"),
        ("10_COST_TRACKING_MASTER", "build_cost_tracking_master"),
    ]

    for sheet_name, builder_func in sheets_to_create:
        ws = wb.create_sheet(sheet_name)

        if builder_func == "basic":
            # Add title and placeholder
            ws['A1'] = sheet_name.replace('_', ' ').upper()
            ws['A1'].font = styles['title']
            ws['A1'].fill = styles['title_fill']
            ws['A1'].alignment = styles['center']
            ws.merge_cells('A1:D1')
            ws.row_dimensions[1].height = 25
            ws['A3'] = "[Content to be populated]"
            ws.column_dimensions['A'].width = 30
        elif builder_func == "build_0_contents":
            build_0_contents(ws, styles)
        elif builder_func == "build_1_input_control":
            build_1_input_control(ws, styles)
        elif builder_func == "build_reference_sheets":
            build_reference_sheets(ws, styles, sheet_name)
        elif builder_func == "build_scenarios_sheet":
            build_scenarios_sheet(ws, styles)
        elif builder_func == "build_cost_tracking_master":
            build_cost_tracking_master(ws, styles)

    # Save workbook
    wb.save(OUTPUT_PATH)
    print(f"✅ Workbook created: {OUTPUT_PATH}")
    print(f"📊 Total sheets: {len(wb.sheetnames)}")
    print(f"📝 Sheet names:")
    for i, sheet in enumerate(wb.sheetnames, 1):
        print(f"   {i:2}. {sheet}")

if __name__ == "__main__":
    build_workbook()
